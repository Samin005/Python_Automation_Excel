import openpyxl as xl
from pathlib import Path


def get_cell_value(row, column):
    # to access a cell
    # cell = sheet['a1']
    # cell = sheet.cell(row=1, column=1)

    # to get cell value
    # cell_value = cell.value
    return sheet.cell(row, column).value


def set_cell_value(row, column, value):
    sheet.cell(row, column).value = value
    print(f'Cell{row, column} updated!')


def save_as(file_name):
    workbook_file.save(file_name)
    print(f'File saved as {file_name}')


def view_existing_files():
    path = Path()
    print('Existing Files:')
    for file in path.glob('*'):
        print(file)
    print('-------')


view_existing_files()
excel_file_name = 'transactions.xlsx'
sheet_name = 'Sheet1'
print(f'Target Excel File: {excel_file_name}')

workbook_file = xl.load_workbook(excel_file_name)
sheet = workbook_file[sheet_name]

header_row_no = 1
row_start_count_without_header = header_row_no + 1
target_column_to_read = 3
target_column_to_update = 4
target_column_header = f'90% of {sheet.cell(1, target_column_to_read).value}'

# setting header of target column
set_cell_value(header_row_no, target_column_to_update, target_column_header)

# setting all values for target column except header
for current_row in range(row_start_count_without_header, sheet.max_row + 1):
    cell_value = get_cell_value(current_row, target_column_to_read)
    set_cell_value(current_row, target_column_to_update, cell_value * 0.9)
save_as('updated_transactions.xlsx')
