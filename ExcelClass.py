import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path
import random


class Excel:
    def __init__(self, path, file_name, sheet_name):
        self.path = path
        self.file_name = file_name
        self.sheet_name = sheet_name
        print(self.path.name)
        if self.path.name != '':
            self.workbook_file = xl.load_workbook(f'{self.path.name}/{self.file_name}')
        else:
            self.workbook_file = xl.load_workbook(self.file_name)
        self.sheet = self.workbook_file[sheet_name]

    def get_cell_value(self, row, column):
        # to access a cell
        # cell = sheet['a1']
        # cell = sheet.cell(row=1, column=1)

        # to get cell value
        # cell_value = cell.value
        return self.sheet.cell(row, column).value

    def set_cell_value(self, row, column, value):
        self.sheet.cell(row, column).value = value
        print(f'Cell{row, column} updated!')

    def create_bar_chart(self, min_row, max_row, min_column, max_column, insert_cell):
        bar_chart = BarChart()
        chart_data = Reference(self.sheet,
                               min_row=min_row,
                               max_row=max_row,
                               min_col=min_column,
                               max_col=max_column)
        bar_chart.add_data(chart_data, titles_from_data=True)
        self.sheet.add_chart(bar_chart, insert_cell)

    def save_as(self, file_name):
        path = Path('results')
        if not path.exists():
            path.mkdir()
        self.workbook_file.save(f'{path.name}/{file_name}')
        print(f'File saved in {path.name}/{file_name}')

    def process_excel_file(self):
        header_row_no = 1
        row_start_count_without_header = header_row_no + 1
        target_column_to_read = 3
        target_column_to_update = 4
        target_column_header = f'90% of {self.sheet.cell(header_row_no, target_column_to_read).value}'

        # setting header of target column
        self.set_cell_value(header_row_no, target_column_to_update, target_column_header)

        # setting all values for target column except header
        for current_row in range(row_start_count_without_header, self.sheet.max_row + 1):
            cell_value = self.get_cell_value(current_row, target_column_to_read)
            self.set_cell_value(current_row, target_column_to_update, cell_value * 0.9)
        self.create_bar_chart(min_row=header_row_no,
                              max_row=self.sheet.max_row,
                              min_column=target_column_to_read,
                              max_column=target_column_to_update,
                              insert_cell='e1')

    def generate_random_integers(self, start_row, end_row, start_column, end_column, min_value, max_value):
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                random_value = random.randint(min_value, max_value)
                self.set_cell_value(row, column, random_value)
